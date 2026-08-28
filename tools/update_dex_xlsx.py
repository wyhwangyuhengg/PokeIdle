# -*- coding: utf-8 -*-
# 根据 src/pokemon-data/pokedex.json 重新生成 docs/宝可梦图鉴.xlsx 数据区：
#   - 数据按全国图鉴顺序，变体紧跟所属本体（0058 → 0058-1 → 0058-2）
#   - legend=true 的行整行红底白字粗体；C3 属性列单属性纯色、双属性左→右线性渐变
#   - 保留原表头、列宽、颜色说明 sheet
import json, shutil
import openpyxl
from copy import copy
from openpyxl.styles import PatternFill, Font, GradientFill
from openpyxl.styles.colors import Color

SRC = r'd:\zc\PokeIdle\src\pokemon-data\pokedex.json'
XLSX = r'd:\zc\PokeIdle\docs\宝可梦图鉴.xlsx'

BERRY_ICONS = ['aspear.png', 'cheri.png', 'chesto.png', 'leppa.png', 'lum.png', 'tamato.png',
               'oran.png', 'pecha.png', 'rawst.png', 'sitrus.png', 'figy.png', 'wiki.png']
BERRY_NAMES = {'aspear.png': '利木果', 'cheri.png': '樱子果', 'chesto.png': '零余果', 'leppa.png': '苹野果',
               'lum.png': '木子果', 'tamato.png': '茄番果', 'oran.png': '橙橙果', 'pecha.png': '桃桃果',
               'rawst.png': '莓莓果', 'sitrus.png': '文柚果', 'figy.png': '勿花果', 'wiki.png': '异奇果'}

# 孵蛋里程参考区间：与游戏 calcHatchDistance 同规则（峰值±σ，按体重/稀有度），舍入到公里
HATCH_MIN, HATCH_MAX = 2000, 30000
def hatch_range(p):
    w = min((p.get('weight') or 100) / 5000, 1)
    r = p.get('rarity') or 0.5
    factor = min(w * 0.6 + r * 0.4, 1)
    mid = HATCH_MIN * (HATCH_MAX / HATCH_MIN) ** factor
    sigma = max(20, mid * 0.2)
    return '%d~%d 公里' % (round((mid - sigma) / 1000), round((mid + sigma) / 1000))

def gender_text(rate):
    if rate == -1:
        return '无性别'
    male, female = (8 - rate) / 8, rate / 8
    if male == 0:
        return '♀100%'
    if female == 0:
        return '♂100%'
    pct = lambda x: '%g%%' % (x * 100)
    return '♂%s ♀%s' % (pct(male), pct(female))

def rows_for(p):
    return [
        p['index'],
        p.get('form') or p['name'],
        '/'.join(p['types']),
        p['region'],
        p['rarity'],
        p['catchRate'],
        gender_text(p.get('genderRate', 4)),
        '、'.join(p.get('eggGroup') or []),
        hatch_range(p),
    ] + list(p['stats']) + [
        '、'.join(BERRY_NAMES[BERRY_ICONS[i]] for i in (p.get('foods') or []))
    ]

shutil.copy(XLSX, XLSX + '.bak_variants')
wb = openpyxl.load_workbook(XLSX)
ws = wb['宝可梦图鉴']
lw = wb['颜色说明']

# 类型 → 颜色（8 位 ARGB）
type_color = {}
for r in range(3, 21):
    t = lw.cell(r, 1).value
    f = lw.cell(r, 1).fill
    if not t or not f:
        continue
    try:
        rgb = f.fgColor.rgb
    except AttributeError:
        continue
    if rgb and str(rgb).startswith('FF'):
        type_color[t] = str(rgb)
print('类型色:', len(type_color), '种')

def make_type_fill(types):
    cols = [type_color.get(t, 'FFA8A77A') for t in types]
    if len(cols) == 1:
        return PatternFill('solid', fgColor=cols[0])
    return GradientFill(degree=0, stop=[Color(rgb=cols[0], tint=0), Color(rgb=cols[1], tint=0)])

# 表头第 1 行在蛋组后插入「孵蛋里程」——原 I~O（HP~爱吃树果）整体右移一列
for c in range(15, 8, -1):
    src = ws.cell(1, c)
    dst = ws.cell(1, c + 1)
    dst.value = src.value
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
h9 = ws.cell(1, 9)
h9.value = '孵蛋里程'

# 通用对齐/边框（取原表第 2 行为模板）
tpl_font = copy(ws.cell(2, 1).font)
tpl_align = copy(ws.cell(2, 1).alignment)
tpl_border = copy(ws.cell(2, 1).border)
tpl_numfmt = ws.cell(2, 1).number_format

NORMAL_FONT = Font(name='宋体', size=10)
LEGEND_FONT = Font(name='宋体', size=10, bold=True, color='FFFFFF')
NORMAL_FILL = PatternFill()
LEGEND_FILL = PatternFill('solid', fgColor='FFF8696B')

# 数据排序：变体紧跟本体，本体按全国图鉴顺序
dex = json.load(open(SRC, encoding='utf-8'))
bases = {}
for p in dex:
    base = p['index'].split('-')[0]
    bases.setdefault(base, []).append(p)
out = []
seen = set()
for p in dex:
    base = p['index'].split('-')[0]
    if base in seen:
        continue
    seen.add(base)
    lst = sorted(bases[base], key=lambda x: (0 if '-' not in x['index'] else 1, x['index']))
    out.extend(lst)
print('总条目:', len(out), '| 变体:', sum(1 for p in out if '-' in p['index']))

# 重建数据区（清空样式，逐行写入）
for r in range(2, 2 + len(out)):
    for c in range(1, 17):
        cell = ws.cell(r, c)
        cell.value = None
        cell.font = copy(tpl_font)
        cell.fill = PatternFill()
        cell.border = copy(tpl_border)
        cell.alignment = copy(tpl_align)
        cell.number_format = tpl_numfmt

for i, p in enumerate(out):
    r = 2 + i
    data = rows_for(p)
    for c, v in enumerate(data, start=1):
        ws.cell(r, c, v)
    type_fill = make_type_fill(p['types'])
    for c in range(1, 17):
        cell = ws.cell(r, c)
        if p.get('legend'):
            cell.font = copy(LEGEND_FONT)
            cell.fill = LEGEND_FILL if c != 3 else type_fill
        elif c == 3:
            cell.fill = type_fill

# 列宽：里程列定宽，末尾多出的食物列沿用原 O 列宽
ws.column_dimensions['I'].width = 12
ws.column_dimensions['P'].width = ws.column_dimensions['O'].width

# 颜色说明表补充说明
lw.cell(32, 1).value = '5. 数据来源：src/pokemon-data/pokedex.json，共 %d 只（本体 %d + 变体 %d）。' % (
    len(out), sum(1 for p in out if '-' not in p['index']), sum(1 for p in out if '-' in p['index']))

wb.save(XLSX)
print('完成：数据行 %d，红色神兽行 %d' % (len(out), sum(1 for p in out if p.get('legend'))))
